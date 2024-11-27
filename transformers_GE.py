import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# Carregar os dados do arquivo
file_path = 'GE 2024 para Tales 1.xlsx'
data = pd.read_excel(file_path)

# Verificar e substituir 'Ground zero' por 'Marco zero' na coluna de classificações
if 'Classificação da última avaliação' in data.columns:
    data['Classificação da última avaliação'] = data['Classificação da última avaliação'].replace('Ground zero', 'Marco zero')
if 'Avaliação de proficiencia (classificação' in data.columns:
    data['Avaliação de proficiencia (classificação'] = data['Avaliação de proficiencia (classificação'].replace('Ground zero', 'Marco zero')

# Verificar se a coluna 'Status' existe antes de filtrar
if 'Status' in data.columns or 'Status ' in data.columns:  # Corrigir possíveis espaços extras na coluna 'Status'
    data.rename(columns=lambda x: x.strip(), inplace=True)  # Remover espaços dos nomes das colunas
    if 'Status' in data.columns:
        # Filtrar apenas os usuários com 'Subsídio ativo'
        data = data[data['Status'] == 'Subsídio ativo']
    else:
        print("A coluna 'Status' não foi encontrada no arquivo Excel. Continuando sem filtrar por 'Subsídio ativo'.")
else:
    print("A coluna 'Status' não foi encontrada no arquivo Excel. Continuando sem filtrar por 'Subsídio ativo'.")

# Ordenar os dados pelo nome do colaborador
if 'Colaborador' in data.columns:
    data = data.sort_values(by='Colaborador')
else:
    raise KeyError("A coluna 'Colaborador' não foi encontrada no arquivo Excel.")

# Lista de classificações da régua da União Europeia (modificada)
idiomas = data['Idiomas'].unique()
tabelas_classificacoes = {}
for idioma in idiomas:
    classificacoes_idioma = data[data['Idiomas'] == idioma]['Classificação da última avaliação'].unique().tolist() + \
                            data[data['Idiomas'] == idioma]['Avaliação de proficiencia (classificação'].unique().tolist()
    tabelas_classificacoes[idioma] = list(set(classificacoes_idioma))  # Remover duplicatas

# Função para ordenar as classificações
def ordenar_classificacoes(classificacoes):
    def classificacao_key(nivel):
        if isinstance(nivel, str) and nivel.lower() == 'marco zero':
            return (0, 0, 0)
        elif isinstance(nivel, str):
            if not '.' in nivel:
                letra = nivel
                return (ord(letra[0]) - ord('A') + int(letra[1])/10, 0, 0)
            nivel = nivel.split('.')
            letra = nivel[0]
            plus = 0
            if '+' in letra:
                letra = letra[:-1]
                plus = 0.1
            numero = int(nivel[1])
            ret = (ord(letra[0]) - ord('A') + int(letra[1])/10, plus, numero)
            return ret

        return (float('inf'), 0, 0)

    return sorted([x for x in classificacoes if (pd.notna(x) and x != "")], key=classificacao_key)

# Função para converter semestre em mês
def semestre_para_mes(semestre_str):
    if isinstance(semestre_str, str):
        match = re.match(r'(\d)[ºo]? Semestre (\d{4})', semestre_str)
        if match:
            semestre_num = int(match.group(1))
            ano = int(match.group(2))
            if semestre_num == 1:
                mes = '01'
            elif semestre_num == 2:
                mes = '07'
            else:
                mes = '01'
            return f'{mes}/{ano}'
        else:
            return semestre_str
    else:
        return semestre_str

# Função para preencher a tabela transformada
def preencher_tabela_transformada(row, classificacoes):
    linha_usuario = {nivel: np.nan for nivel in classificacoes}
    ciclo_origem = row.get('Data da primeira avaliação (proficiencia)', None)
    ciclo_atual = row.get('Data da última avaliação', None)
    prazo_meta = row.get('Validade do subsidio (24 meses)', None)

    if pd.notna(ciclo_origem):
        ciclo_origem = pd.to_datetime(ciclo_origem).strftime('%m/%Y')
    if pd.notna(ciclo_atual):
        ciclo_atual = pd.to_datetime(ciclo_atual).strftime('%m/%Y')

    if pd.notna(prazo_meta):
        prazo_meta = pd.to_datetime(prazo_meta).strftime('%m/%Y')

    # Converter ciclos de semestre para mês
    if ciclo_origem is not None and pd.notna(ciclo_origem):
        ciclo_origem = semestre_para_mes(ciclo_origem)
    if ciclo_atual is not None and pd.notna(ciclo_atual):
        ciclo_atual = semestre_para_mes(ciclo_atual)
    if prazo_meta is not None and pd.notna(prazo_meta):
        prazo_meta = semestre_para_mes(prazo_meta)

    ultima_classificacao = row.get('Classificação da última avaliação', None)
    proficiencia_original = row.get('Avaliação de proficiencia (classificação', None)
    meta_final = row.get('Meta final', None)

    # Preencher ciclo de origem
    if proficiencia_original is not None and pd.notna(proficiencia_original) and proficiencia_original in classificacoes:
        linha_usuario[proficiencia_original] = ciclo_origem

    # Preencher última classificação
    if ultima_classificacao is not None and pd.notna(ultima_classificacao) and ultima_classificacao in classificacoes:
        linha_usuario[ultima_classificacao] = ciclo_atual

    # Preencher meta final
    if meta_final is not None and pd.notna(meta_final) and meta_final in classificacoes:
        linha_usuario[meta_final] = prazo_meta

    # Colorir o caminho percorrido
    if ultima_classificacao is not None and proficiencia_original is not None and \
            ultima_classificacao in classificacoes and proficiencia_original in classificacoes and ciclo_atual is not None:
        classificacoes_indices = [classificacoes.index(ultima_classificacao), classificacoes.index(proficiencia_original)]
        for i in range(min(classificacoes_indices), max(classificacoes_indices) + 1):
            if classificacoes[i] != ultima_classificacao and classificacoes[i] != proficiencia_original:
                linha_usuario[classificacoes[i]] = 'Caminho Percorrido'

    # Adicionar caminho a percorrer entre ciclo_atual e prazo_meta
    if ultima_classificacao is not None and meta_final is not None and \
            ultima_classificacao in classificacoes and meta_final in classificacoes:
        classificacoes_indices = [classificacoes.index(ultima_classificacao), classificacoes.index(meta_final)]

        for i in range(min(classificacoes_indices), max(classificacoes_indices) + 1):
            if classificacoes[i] != ultima_classificacao and classificacoes[i] != meta_final and classificacoes[i] != proficiencia_original:
                linha_usuario[classificacoes[i]] = 'Caminho a Percorrer'

    # Adicionar caminho a percorrer entre ciclo_origem e prazo_meta se ciclo_atual estiver ausente
    if ciclo_atual is None or pd.isna(ciclo_atual):
        if proficiencia_original is not None and meta_final is not None and \
                proficiencia_original in classificacoes and meta_final in classificacoes:
            classificacoes_indices = [classificacoes.index(proficiencia_original), classificacoes.index(meta_final)]

            for i in range(min(classificacoes_indices), max(classificacoes_indices) + 1):
                if classificacoes[i] != proficiencia_original and classificacoes[i] != meta_final:
                    linha_usuario[classificacoes[i]] = 'Caminho a Percorrer'

    return pd.Series(linha_usuario)

# Deletar o arquivo existente, se houver
output_file = 'Tabela_Transformada_GE_Tales.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)

# Criar tabelas separadas por idioma e salvar em Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for idioma in idiomas:
        classificacoes = ordenar_classificacoes(tabelas_classificacoes[idioma])
        data_filtrada = data[data['Idiomas'] == idioma]
        tabela_transformada = data_filtrada.apply(preencher_tabela_transformada, axis=1, classificacoes=classificacoes)
        usuarios = data_filtrada['Colaborador']
        tabela_transformada.index = usuarios
        tabela_transformada.to_excel(writer, sheet_name=idioma)

        # Ajustar tamanhos das células e aplicar cores
        workbook = writer.book
        worksheet = writer.sheets[idioma]

        # Garantir que a planilha esteja visível
        worksheet.sheet_state = 'visible'

        # Ajustar largura das colunas
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Coluna
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Aplicar cores ao cabeçalho e células
        header_fill = PatternFill(start_color='1f2f36', end_color='1f2f36', fill_type='solid')
        cell_fill = PatternFill(start_color='391e70', end_color='391e70', fill_type='solid')
        cell_fill_meta_final = PatternFill(start_color='adc22f', end_color='adc22f', fill_type='solid')
        cell_fill_caminho_a_percorrer = PatternFill(start_color='adc22f', end_color='adc22f', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(color='FFFFFF')

        # Colorir cabeçalho
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.font = header_font  # Texto branco
            cell.border = thin_border

        # Colorir primeira coluna
        for cell in worksheet['A']:
            if cell.row > 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Colorir células preenchidas, o caminho percorrido, e aplicar bordas e alinhamento
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            aux = 0
            for cell in row:
                if cell.value not in ['Caminho Percorrido', 'Caminho a Percorrer', ""]:
                    aux += 1

            aux = 0
            for cell in row:
                if cell.value == 'Caminho Percorrido':
                    cell.fill = cell_fill
                    cell.value = ""
                elif cell.value == 'Caminho a Percorrer':
                    cell.fill = cell_fill_caminho_a_percorrer
                    cell.value = ""
                    aux = 2
                elif cell.value is not None and cell.value != "":
                    cell.fill = cell_fill

                    aux += 1
                    if aux == 3:
                        cell.fill = cell_fill_meta_final
                    else:
                        cell.font = header_font
                    cell.border = thin_border
                cell.alignment = cell_alignment

        # Fixar primeira linha e primeira coluna
        worksheet.freeze_panes = 'B2'

print("Tabela salva com sucesso!")
